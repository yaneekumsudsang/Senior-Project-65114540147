from datetime import datetime
from django.conf import settings
from django.contrib.admin import options
from Kouponapp.models import *
import os
from django.apps import apps
from django.db.models import *
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from Kouponapp.models import Member, Store
from decimal import Decimal
import traceback
from decimal import Decimal
from datetime import datetime
from django.core.exceptions import ValidationError
from django.db import IntegrityError

class Command(BaseCommand):
    help = "Load promotion data from data7.xlsx file"

    def handle(self, *args, **kwargs):
        # Path ของไฟล์ Excel
        #file_path = '/Users/yaneekumsudsang/Koupon/data4.xlsx'
        file_path = os.path.join(settings.BASE_DIR, 'Kouponapp/fixtures/data7.xlsx')

        # Load Excel workbook
        wb = load_workbook(filename=file_path)

        wm = wb['Member']
        for row in wm:
            values = [cell.value for cell in row]
            if values[0] != 'id':
                #id = values[0]
                #username = values[1]
                user = User.objects.create_user(values[1], values[5], str(values[6]), pk=values[0], first_name=values[2], last_name=values[3])
                member, created = Member.objects.get_or_create(user=user, phone=values[4], is_owner=bool(values[8]))

        store_sheet = wb['Store']
        self.stdout.write("Loading Store Data...")

        for row in store_sheet.iter_rows(min_row=2, values_only=True):
            store_id, store_name, owner_id = row

            # ตรวจสอบข้อมูลที่จำเป็น
            if not store_id or not store_name or not owner_id:
                self.stdout.write(f"Skipping row due to missing data: {row}")
                continue

            try:
                # ค้นหา Member ที่เป็นเจ้าของร้าน
                member = Member.objects.get(id=owner_id, is_owner=True)
            except Member.DoesNotExist:
                self.stdout.write(
                    f"Owner ID {owner_id} not found or not marked as owner. Skipping store: {store_name}.")
                continue

            try:
                # สร้างหรืออัปเดตร้านค้า
                store_instance, created = Store.objects.update_or_create(
                    id=store_id,
                    defaults={'store_name': store_name, 'owner': member}
                )

                if created:
                    self.stdout.write(self.style.SUCCESS(f"Created store: {store_name} (Owner ID: {owner_id})"))
                else:
                    self.stdout.write(self.style.WARNING(f"Updated store: {store_name} (Owner ID: {owner_id})"))

            except Exception as e:
                self.stdout.write(self.style.ERROR(f"Error processing store {store_name}: {str(e)}"))

        #Promotion
        promotion_sheet = wb['Promotion']
        for row in promotion_sheet.iter_rows(min_row=2, values_only=True):
            promotion_id, store_id, picture, cupsize, cups, discount, free, name, details, start, end, count, *_ = row

            # ตรวจสอบค่า store_id เป็น int
            if not isinstance(store_id, int):
                self.stdout.write(f"Invalid store_id: {store_id}. Skipping promotion {name}.")
                continue

            store = Store.objects.filter(id=store_id).first()
            if not store:
                self.stdout.write(f"Store ID {store_id} not found. Skipping promotion {name}.")
                continue

            try:
                start = datetime.strptime(str(start).split()[0], '%Y-%m-%d').date()
                end = datetime.strptime(str(end).split()[0], '%Y-%m-%d').date()
            except ValueError as e:
                self.stdout.write(f"Invalid date format for promotion ID {promotion_id}. Skipping... Error: {e}")
                continue

            count = int(str(count).strip()) if str(count).strip().isdigit() else 0

            promotion_instance, created = Promotion.objects.update_or_create(
                store_id=store_id,
                name=name,
                defaults={
                    'picture': picture,
                    'cupsize': cupsize,
                    'cups': cups,
                    'discount': discount,
                    'free': free,
                    'details': details,
                    'start': start,
                    'end': end,
                    'count': count,
                }
            )

            if created:
                self.stdout.write(f"Created promotion: {name} for store: {store.store_name} with count: {count}")
            else:
                self.stdout.write(f"Updated promotion: {name} with count: {count}")

        # โหลดข้อมูลคูปองจากชีต Excel
        coupon_sheet = wb['Coupon']
        for row in coupon_sheet.iter_rows(min_row=2, values_only=True):
            coupon_id, promotion_id, promotion_count, collect, member_id, collect_qr_code_url, use_qr_code_url, *_ = row

            # ค้นหา Promotion
            promotion = Promotion.objects.filter(id=promotion_id).first()
            if not promotion:
                self.stdout.write(f"Promotion ID {promotion_id} not found. Skipping coupon ID {coupon_id}.")
                continue

            # ค้นหาคูปองที่มีอยู่แล้วในฐานข้อมูล
            coupon = Coupon.objects.filter(promotion=promotion, promotion_count=promotion_count).first()

            # กำหนด username ของเจ้าของร้าน
            store_owner = getattr(promotion.store.owner, 'user', None)
            username = store_owner.username if store_owner else 'unknown'

            # กำหนดค่า QR Code URL
            generated_collect_qr_code_url = f"http://127.0.0.1/koupon/qr/{promotion.id}/collec/{promotion_count}/{coupon_id}/"
            generated_use_qr_code_url = f"http://127.0.0.1/koupon/qr/{promotion.id}/use/{promotion_count}/{coupon_id}/"

            if not coupon:
                # คูปองยังไม่มีในระบบ ให้สร้างใหม่
                coupon = Coupon.objects.create(
                    promotion=promotion,
                    promotion_count=promotion_count,
                    collect=bool(collect),
                    member_id=member_id if member_id else None,
                    collect_qr_code_url=generated_collect_qr_code_url,
                    use_qr_code_url=generated_use_qr_code_url
                )
                self.stdout.write(
                    f"Created new coupon ID {coupon.id} for promotion {promotion.id} (count {promotion_count})")
            else:
                # อัปเดตข้อมูลของคูปองให้ตรงกับไฟล์ Excel
                updated_fields = {}
                if coupon.collect_qr_code_url != generated_collect_qr_code_url:
                    updated_fields['collect_qr_code_url'] = generated_collect_qr_code_url
                if coupon.use_qr_code_url != generated_use_qr_code_url:
                    updated_fields['use_qr_code_url'] = generated_use_qr_code_url
                if coupon.collect != bool(collect):
                    updated_fields['collect'] = bool(collect)
                if coupon.member_id != (member_id if member_id else None):
                    updated_fields['member_id'] = member_id if member_id else None

                if updated_fields:
                    for field, value in updated_fields.items():
                        setattr(coupon, field, value)
                    coupon.save()
                    self.stdout.write(
                        f"Updated coupon ID {coupon.id} for promotion {promotion.id} (count {promotion_count})")
                else:
                    self.stdout.write(f"No updates required for coupon ID {coupon.id}")

        self.stdout.write(self.style.SUCCESS("Data loaded successfully!"))
